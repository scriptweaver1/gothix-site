#!/usr/bin/env python3
"""
GothicMynx masterlist build script.

Reads the Reddit and Patreon spreadsheets, normalizes durations, merges
cross-posted audios (same title on both platforms → one card with both
links), and writes a single audios.json the website consumes at runtime.

Usage:
    python build.py

Inputs  (same folder):  reddit_sorted.xlsx , patreon_sorted.xlsx
Output  (same folder):  audios.json
"""

import json
import re
from datetime import time, timedelta, datetime, date
from openpyxl import load_workbook

REDDIT_XLSX  = "reddit_sorted.xlsx"
PATREON_XLSX = "patreon_sorted.xlsx"
OUTPUT_JSON  = "audios.json"

RED = "FFFF0000"


# ----------------------------------------------------------------------
# Value normalizers
# ----------------------------------------------------------------------

def parse_duration(v):
    """Return a clean 'M:SS' string from the sheet's mangled duration value.

    Google Sheets stores these as:
      - time  value shown as hh:mm  -> digits are really MM:SS
      - timedelta shown as [hh]:mm:ss (seconds 0) -> hh:mm are really MM:SS
      - occasional string like '32:17:00.000'
    """
    if v is None or v == "":
        return ""
    if isinstance(v, str):
        parts = v.strip().split(":")
        if len(parts) >= 2:
            try:
                mm = int(float(parts[0])); ss = int(float(parts[1]))
                return f"{mm}:{ss:02d}"
            except ValueError:
                return v.strip()
        return v.strip()
    if isinstance(v, time):
        return f"{v.hour}:{v.minute:02d}"
    if isinstance(v, timedelta):
        total = int(round(v.total_seconds()))
        return f"{total // 3600}:{(total % 3600) // 60:02d}"
    if isinstance(v, datetime):
        return f"{v.hour}:{v.minute:02d}"
    return str(v)


def to_iso(v):
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip().split(" ")[0]


def split_tags(raw):
    if not raw:
        return []
    s = str(raw).strip()
    # Bracketed format: [Tag one][Tag two][Tag three]  (also handles ]to[ connectors)
    brackets = re.findall(r"\[([^\]]+)\]", s)
    if brackets:
        parts = brackets
    else:
        # Fallback: comma / semicolon separated
        parts = str(s).replace(";", ",").split(",")
    out, seen = [], set()
    for p in parts:
        t = p.strip().strip(",").strip()
        if t and t.lower() not in seen:
            out.append(t); seen.add(t.lower())
    return out


def norm_key(title):
    t = (title or "").strip().lower()
    t = re.sub(r"\s+", " ", t)
    return t.strip(" .!?,;:\"'")


def persona_label(v):
    s = str(v).strip()
    if s in ("1", "1.0"):
        return "Gothix"
    if s in ("0", "0.0"):
        return "Minxy"
    return None


def is_red(ws, row):
    for c in (1, 3):
        cell = ws.cell(row=row, column=c)
        if cell.fill and cell.fill.patternType == "solid" and cell.fill.fgColor.rgb == RED:
            return True
    return False


# ----------------------------------------------------------------------
# Sheet reader
# ----------------------------------------------------------------------

def read_sheet(path, source):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h not in (None, ""):
            headers[str(h).strip()] = c

    def cell(row, name):
        col = headers.get(name)
        return ws.cell(row=row, column=col).value if col else None

    link_col = "Reddit Link" if source == "reddit" else "Patreon Link"
    rows = []
    for row in range(2, ws.max_row + 1):
        title = cell(row, "Title")
        if title in (None, ""):
            continue
        link = cell(row, link_col)
        entry = {
            "title": str(title).strip(),
            "persona": persona_label(cell(row, "1- Gothix/Mynx - 0")),
            "category": (str(cell(row, "Category")).strip() if cell(row, "Category") else ""),
            "audience": (str(cell(row, "Audience")).strip() if cell(row, "Audience") else ""),
            "tags": split_tags(cell(row, "Tags")),
            "date": to_iso(cell(row, "Date")),
            "duration": parse_duration(cell(row, "Duration")),
            "synopsis": (str(cell(row, "Synopsis")).strip() if cell(row, "Synopsis") else ""),
            "writer": (str(cell(row, "Writer")).strip() if cell(row, "Writer") else ""),
            "scriptLink": (str(cell(row, "Script link")).strip() if cell(row, "Script link") else ""),
            "redditLink": str(link).strip() if (source == "reddit" and link) else "",
            "patreonLink": str(link).strip() if (source == "patreon" and link) else "",
            "noViableLinks": is_red(ws, row),
        }
        # Reddit-only extras
        if source == "reddit":
            cp = cell(row, "Collab Partner 1")
            ed = cell(row, "Editor")
            entry["collabPartner"] = str(cp).strip() if cp else ""
            entry["editor"] = str(ed).strip() if ed else ""
        rows.append(entry)
    return rows


# ----------------------------------------------------------------------
# Merge (cross-post dedup: same title on both platforms -> one card)
# ----------------------------------------------------------------------

CLEAN_PLACEHOLDERS = {"", "n/a", "na", "none", "-"}


def clean(val):
    return "" if str(val).strip().lower() in CLEAN_PLACEHOLDERS else str(val).strip()


def merge_into(base, extra):
    """Fold `extra` into `base` (base = Reddit-priority primary)."""
    base["redditLink"]  = base["redditLink"]  or extra.get("redditLink", "")
    base["patreonLink"] = base["patreonLink"] or extra.get("patreonLink", "")
    for f in ("duration", "synopsis", "writer", "category", "audience", "persona"):
        if not clean(base.get(f)) and clean(extra.get(f)):
            base[f] = extra[f]
    # union of tags, preserving order
    seen = {t.lower() for t in base["tags"]}
    for t in extra.get("tags", []):
        if t.lower() not in seen:
            base["tags"].append(t); seen.add(t.lower())
    # a card is only "no viable links" if BOTH sources agree it's dead
    base["noViableLinks"] = base.get("noViableLinks") and extra.get("noViableLinks", False)


def build():
    reddit  = read_sheet(REDDIT_XLSX,  "reddit")
    patreon = read_sheet(PATREON_XLSX, "patreon")

    merged = {}     # norm_key -> entry (insertion order = reddit first)
    order  = []

    def add(entry):
        k = norm_key(entry["title"])
        if k in merged:
            merge_into(merged[k], entry)
        else:
            merged[k] = entry
            order.append(k)

    for e in reddit:
        add(e)
    for e in patreon:
        add(e)

    audios = []
    for k in order:
        e = merged[k]
        # Skip entries with nowhere to link
        if not e["redditLink"] and not e["patreonLink"]:
            continue

        obj = {"title": e["title"]}
        if e["persona"]:  obj["persona"] = e["persona"]
        if clean(e["category"]): obj["category"] = e["category"]
        if clean(e["audience"]): obj["audience"] = e["audience"]
        # audience as leading tag, de-duplicated
        tags = list(e["tags"])
        aud = clean(e["audience"])
        if aud and aud.lower() not in [t.lower() for t in tags]:
            tags.insert(0, aud)
        if tags: obj["tags"] = tags
        if e["date"]: obj["date"] = e["date"]
        if clean(e["duration"]): obj["duration"] = e["duration"]
        if clean(e["synopsis"]): obj["synopsis"] = e["synopsis"]
        if clean(e["writer"]): obj["writer"] = e["writer"]
        if clean(e["scriptLink"]) and clean(e["scriptLink"]).lower().startswith("http"):
            obj["scriptLink"] = e["scriptLink"]
        if e["redditLink"]: obj["redditLink"] = e["redditLink"]
        if e["patreonLink"]: obj["patreonLink"] = e["patreonLink"]
        if clean(e.get("collabPartner", "")): obj["collabPartner"] = e["collabPartner"]
        if clean(e.get("editor", "")): obj["editor"] = e["editor"]
        if e["noViableLinks"]: obj["noViableLinks"] = True
        audios.append(obj)

    # Newest first (blank dates sink to the bottom)
    audios.sort(key=lambda a: a.get("date", ""), reverse=True)

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(audios, f, ensure_ascii=False, indent=2)

    # Report
    both = sum(1 for a in audios if a.get("redditLink") and a.get("patreonLink"))
    r_only = sum(1 for a in audios if a.get("redditLink") and not a.get("patreonLink"))
    p_only = sum(1 for a in audios if a.get("patreonLink") and not a.get("redditLink"))
    dead = sum(1 for a in audios if a.get("noViableLinks"))
    cats = sorted({a.get("category", "Uncategorized") for a in audios})
    print(f"Reddit rows: {len(reddit)}  |  Patreon rows: {len(patreon)}")
    print(f"Merged unique audios: {len(audios)}")
    print(f"  both platforms: {both}   reddit-only: {r_only}   patreon-only: {p_only}")
    print(f"  flagged no-viable-links: {dead}")
    print(f"  categories: {cats}")
    print(f"Wrote {OUTPUT_JSON}")


if __name__ == "__main__":
    build()
